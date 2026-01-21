import os
import json
import smtplib
from email.message import EmailMessage
from datetime import datetime, timezone
from typing import Dict, Optional

from google import genai
from google.genai import types
from openpyxl import Workbook, load_workbook


# =====================
# CONFIG
# =====================
MODEL = "gemini-2.5-flash"
XLSX_PATH = "prompts_log.xlsx"

SHEET_CRM = "CRM"
SHEET_SUPPORT = "Support Tickets"

SMTP_TO_DEFAULT = "hayk_nalchajyan@edu.aua.am"

SYSTEM_INSTRUCTIONS = (
    "You are an automated customer inquiry triage system.\n"
    "Return ONLY valid JSON (no markdown, no extra text).\n"
    "JSON must contain EXACTLY these keys:\n"
    "sentiment, intent, urgency_score, summary, suggested_reply.\n"
    "Allowed values:\n"
    "- sentiment: Positive | Neutral | Negative\n"
    "- intent: Sales | Support | Spam | Other\n"
    "- urgency_score: integer 1..10\n"
    "summary MUST be in Armenian."
)

REQUIRED_KEYS = ["sentiment", "intent", "urgency_score", "summary", "suggested_reply"]
ALLOWED_INTENT = {"Sales", "Support", "Spam", "Other"}

HEADERS = [
    "timestamp_utc",
    "customer_message",
    "sentiment",
    "intent",
    "urgency_score",
    "summary_hy",
    "suggested_reply",
    "lead_tag",
    "email_status",
    "email_error",
]


# =====================
# EMAIL (GMAIL SMTP)
# =====================
def send_email_hot_lead(to_email: str, subject: str, body: str) -> None:
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")

    if not smtp_user or not smtp_pass:
        raise RuntimeError("SMTP_USER / SMTP_PASS missing (Gmail App Password required)")

    msg = EmailMessage()
    msg["From"] = smtp_user
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body)

    with smtplib.SMTP("smtp.gmail.com", 587, timeout=20) as server:
        server.ehlo()
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)


# =====================
# EXCEL HELPERS
# =====================
def _ensure_headers(ws) -> None:
    """If the first row doesn't match headers, write headers into row 1."""
    first_row = [ws.cell(row=1, column=i + 1).value for i in range(len(HEADERS))]
    if first_row != HEADERS:
        # If sheet is empty -> append headers
        if ws.max_row == 1 and ws["A1"].value is None:
            ws.append(HEADERS)
        else:
            # overwrite row 1 with headers (keeps existing data below)
            for i, h in enumerate(HEADERS, start=1):
                ws.cell(row=1, column=i).value = h


def ensure_workbook(path: str) -> None:
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        if wb.active.title == "Sheet":
            wb.remove(wb.active)

    if SHEET_CRM not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_CRM)
        ws.append(HEADERS)
    else:
        _ensure_headers(wb[SHEET_CRM])

    if SHEET_SUPPORT not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_SUPPORT)
        ws.append(HEADERS)
    else:
        _ensure_headers(wb[SHEET_SUPPORT])

    wb.save(path)


def append_row(path: str, sheet: str, row: list) -> None:
    wb = load_workbook(path)
    ws = wb[sheet]
    ws.append(row)
    wb.save(path)


# =====================
# GEMINI
# =====================
def call_gemini(prompt: str) -> Dict:
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY missing")

    client = genai.Client(api_key=api_key)
    resp = client.models.generate_content(
        model=MODEL,
        contents=prompt,
        config=types.GenerateContentConfig(
            system_instruction=SYSTEM_INSTRUCTIONS,
            response_mime_type="application/json",
            temperature=0.2,
        ),
    )

    # Make JSON errors explicit (so you see them in n8n)
    try:
        data = json.loads(resp.text)
    except Exception as e:
        raise ValueError(f"Model did not return valid JSON: {e}. Raw: {resp.text[:500]}")

    # Minimal validation (prevents key errors later)
    for k in REQUIRED_KEYS:
        if k not in data:
            raise ValueError(f"Missing key '{k}' in model JSON. Raw: {resp.text[:500]}")
    if data.get("intent") not in ALLOWED_INTENT:
        raise ValueError(f"Invalid intent '{data.get('intent')}'. Raw: {resp.text[:500]}")

    return data


# =====================
# ROUTER
# =====================
def run_pipeline(customer_message: str, hot_lead_email_to: Optional[str] = None) -> Dict:
    ensure_workbook(XLSX_PATH)

    prompt = f"""
Analyze the customer message and output ONLY JSON in this exact format:

{{
  "sentiment": "Positive|Neutral|Negative",
  "intent": "Sales|Support|Spam|Other",
  "urgency_score": 1-10,
  "summary": "Հաղորդագրության հակիրճ նկարագրություն՝ հայերեն",
  "suggested_reply": "AI-ի կողմից գեներացված պատասխանի սևագիր"
}}

Customer message:
{customer_message}
""".strip()

    ts = datetime.now(timezone.utc).isoformat()
    parsed = call_gemini(prompt)

    intent = parsed["intent"]
    urgency = int(parsed["urgency_score"])

    # Attach routing metadata (for n8n visibility)
    parsed["routed_case"] = ""
    parsed["lead_tag"] = ""
    parsed["email_status"] = ""
    parsed["email_error"] = ""

    # ---- SPAM: END ----
    if intent == "Spam":
        parsed["routed_case"] = "SPAM_NO_ACTION"
        print("→ Routed as SPAM (no action)")
        return parsed

    row = [
        ts,
        customer_message,
        parsed["sentiment"],
        intent,
        urgency,
        parsed["summary"],
        parsed["suggested_reply"],
        "",  # lead_tag
        "",  # email_status
        "",  # email_error
    ]

    # ---- SALES ----
    if intent == "Sales":
        parsed["routed_case"] = "SALES"
        if urgency > 7:
            parsed["routed_case"] = "SALES_HOT_LEAD"
            parsed["lead_tag"] = "HOT LEAD"

            try:
                to_addr = hot_lead_email_to or SMTP_TO_DEFAULT
                send_email_hot_lead(
                    to_addr,
                    subject="HOT LEAD",
                    body=f"""HOT LEAD DETECTED

Urgency: {urgency}

Customer message:
{customer_message}

Summary (hy):
{parsed["summary"]}

Suggested reply:
{parsed["suggested_reply"]}
""",
                )
                parsed["email_status"] = "SENT"
            except Exception as e:
                parsed["email_status"] = "FAILED"
                parsed["email_error"] = str(e)

        # write CRM row always for Sales
        row[7] = parsed["lead_tag"]
        row[8] = parsed["email_status"]
        row[9] = parsed["email_error"]
        append_row(XLSX_PATH, SHEET_CRM, row)
        print("→ Routed to CRM")
        return parsed

    # ---- SUPPORT / OTHER ----
    parsed["routed_case"] = "SUPPORT" if intent == "Support" else "OTHER"
    append_row(XLSX_PATH, SHEET_SUPPORT, row)
    print("→ Routed to Support Tickets")
    return parsed


if __name__ == "__main__":
    print("XLSX location:", os.path.abspath(XLSX_PATH))

    examples = {
        "HOT_LEAD": (
            "Hello, we already have management approval for the Enterprise plan "
            "and a confirmed budget of $2,000/month. We need to sign the contract "
            "today or tomorrow at the latest. Please send pricing and SLA ASAP."
        ),
        "SUPPORT": (
            "Hi, our API requests to /v1/reports started returning 500 errors "
            "since this morning. This blocks our production workflow. Please investigate."
        ),
        "SPAM": (
            "Congratulations! You won a free iPhone. Click this link and enter "
            "your bank details to receive the prize today!"
        ),
    }

    for name, message in examples.items():
        print(f"\n--- RUNNING EXAMPLE: {name} ---")
        result = run_pipeline(message)
        print(json.dumps(result, ensure_ascii=False, indent=2))
